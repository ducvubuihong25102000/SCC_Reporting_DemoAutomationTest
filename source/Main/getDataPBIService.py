from time import time
start = time()

import openpyxl
import requests
import json
import pandas as pd
from source.Connections.connPBIService import headers
from source.Connections.Credentials import excel_quality, EXCEL_PATH

# # Get Rest API list of workspace
# request_list_workspace = 'https://api.powerbi.com/v1.0/myorg/groups'

# response_list_workspace = requests.get(request_list_workspace, headers= headers)
# print(response_list_workspace.status_code)
# print(response_list_workspace.json())
# print('--------------------------------------------------')

# # Check permission
# if response_list_workspace.status_code == 200:
#     result = response_list_workspace.json()5
#     for item in result['value']:
#         # print(item)
#         # Get list of dataset in each workspace
#         request_list_datasets = f'https://api.powerbi.com/v1.0/myorg/groups/{item["id"]}/datasets'
#         print(request_list_datasets)
#         response_list_datasets = requests.get(request_list_datasets, headers= headers)
#         print(item['name'])
#         print(item['id'])
#         print(response_list_datasets.status_code)
#         for item in response_list_datasets.json()['value']:
#             print(item['id']) 
#         print('--------------------------------------------------')
# else:
#     print('Refresh User Permissions')
#     requests.post('https://api.powerbi.com/v1.0/myorg/RefreshUserPermissions', headers= headers)
#     new_response_request = requests.get(request_list_workspace, headers= headers)
#     print(new_response_request.status_code) 
#     new_result = new_response_request.json()
#     for new_item in new_result['value']:
#         print(new_item)
        
        

# # DAX query (SELECT * FROM Customer)
# DAX_query = {
#     'queries' : [
#         {
#             'query': r'EVALUATE VALUES(Customer)'
#         }
#     ],
#     "serializerSettings": {
#     "includeNulls": True
#     } 
# }
# datasetId = '753c0d56-2e7e-42f0-a29c-3648a45c022d' # Get from above API
# request_datasets_query = f'https://api.powerbi.com/v1.0/myorg/datasets/{datasetId}/executeQueries'

# response_datasets_query = requests.post(request_datasets_query, headers= headers, json= DAX_query)
# print('Request Power BI Status Code: ' + str(response_datasets_query.status_code))

# result_json = response_datasets_query.json()
# # clean_response_datasets_query = json.dumps(result_json, indent= 2)
# # print(clean_response_datasets_query)

# df_pbi = pd.DataFrame.from_dict(result_json['results'][0]['tables'][0]['rows'])
# print(df_pbi)
# print('Time after call API: ' + str(time() - start))

# myworkbook = openpyxl.load_workbook(r'D:\DemoPowerBI\demo_powerbi\source\PythonDataQualityChecker.xlsx')
# worksheet = myworkbook['Quality Check']
# cell = worksheet.cell(row= 3, column= 3, value= str(df_pbi.values))
# myworkbook.save(r'D:\DemoPowerBI\demo_powerbi\source\PythonDataQualityChecker.xlsx')


#DAX Query (SELECT Id_Cust, SUM(Quantity) FROM Order ORDER BY Quantity, Id_Cust)
DAX_query = {
    'queries' : [
        {
            'query' : ''
        }
    ],
    "serializerSettings": {
    "includeNulls": True
    } 
}
DAX_query['queries'][0]['query'] = excel_quality['PBI_DAX_Query'][0]

datasetId = '753c0d56-2e7e-42f0-a29c-3648a45c022d' # Get from above API
request_datasets_query = f'https://api.powerbi.com/v1.0/myorg/datasets/{datasetId}/executeQueries'

response_datasets_query = requests.post(request_datasets_query, headers= headers, json= DAX_query)
print('Request Power BI Status Code: ' + str(response_datasets_query.status_code))

result_json = response_datasets_query.json()
# clean_response_datasets_query = json.dumps(result_json, indent= 2)
# print(clean_response_datasets_query)

df_pbi = pd.DataFrame.from_dict(result_json['results'][0]['tables'][0]['rows']).sort_values(by=['[Quantity]', 'Order[Id_Cust]'])


myworkbook = openpyxl.load_workbook(EXCEL_PATH)
worksheet = myworkbook['Quality Check']
cell = worksheet.cell(row= 2, column= 3, value= str(df_pbi.values))
myworkbook.save(EXCEL_PATH)

print(df_pbi)
print('Time after call API: ' + str(time() - start))