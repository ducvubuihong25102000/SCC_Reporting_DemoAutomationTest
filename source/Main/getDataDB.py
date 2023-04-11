# Solution add specific path by system path
# import sys
# sys.path.insert(0,'D:\\DemoPowerBI\\demo_powerbi\\source\\') #Add path for Connections package
# sys.path.insert(1, 'D:\\DemoPowerBI\\demo_powerbi\\source\\Connections\\') #Add path for Cendential package
# import xlswriter
import openpyxl
import pandas as pd
from time import time
from source.Connections import connDB as conn
from source.Connections.Credentials import excel_quality, EXCEL_PATH

query_cust = 'SELECT * FROM dbo.Customer'
query_pro = 'SELECT Id_Pro FROM dbo.Product'

cursor = conn.returnConn[0]

start = time()
# list_customer = [tuple(i) for i in cursor.execute(query_cust).fetchall()]
# list_product_id = [i[0] for i in cursor.execute(query_pro).fetchall()]

# df = pd.DataFrame(list_customer, columns= ['Customer[Id_Cust]', 'Customer[CustName]', 'Customer[CustPhone]'])
# print(df) 
# myworkbook = openpyxl.load_workbook(EXCEL_PATH)
# worksheet = myworkbook.get_sheet_by_name('Quality Check')
# cell = worksheet.cell(row= 3, column= 4, value= str(df.values))
# myworkbook.save(EXCEL_PATH)
# print('Time after query DB: '+ str(time()-start)) #0.036499738693237305


query_sum_aggregate = excel_quality['SQL_Query'][0]
print(query_sum_aggregate)
result = [tuple(i) for i in cursor.execute(query_sum_aggregate).fetchall()]
df = pd.DataFrame(result, columns= ['Order[Id_Cust]', '[Quantity]'])


myworkbook = openpyxl.load_workbook(EXCEL_PATH)
worksheet = myworkbook.get_sheet_by_name('Quality Check')
cell = worksheet.cell(row= 2, column= 4, value= str(df.values))
myworkbook.save(EXCEL_PATH)

print(df) 
print('Time after query DB: '+ str(time()-start))
